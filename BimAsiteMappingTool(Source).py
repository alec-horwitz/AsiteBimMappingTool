'''*****************************************************************************
Name: BimAsiteMappingTool

Purpose: This program provides analytical information about the Compny IDs, 
	 Names, and Statuses from the BIM and Asite databases for 
	 prequalification to help make informed decisions aboout how best to 
	 maintain and synchronize the information on the two databases.

Algorithm: Gets the Compny IDs, Names, and Statuses from the BIM and Asite
           directories. Reads in a Bim to Asite status translation file to
	   properly compare the statuses of the two data bases. Checks to see if
	   all IDs in Asite exsit in BIM and if all IDs in BIM exsit in Asite. 
	   If a BIM company does not have an ID it will be flaged as missing 
	   from Asite. For each company that the ID exsist on both it checks if
	   the Names and/or Statuses match on both. It the spits out the data on
	   a spreadsheet.

Input: A file called settings.txt that determins the names/locations of another 
       .txt file(for the status translations), a .csv file(containing the 
       required data) and a .xlsx file(containing the other required data) which
       also then get read in.
       
Output: A .xlsx file with the IDs, Names, and Statuses of each company as well as
        which data base each company is missing from if either. For each company 
	that is not missing from either, the same file will cantain wheather 
	there is a difference in the status and how the company name is spelled 
	across the two databases.

Author: Alec Horwitz
Email: horwitz_alec@wheatoncollege.edu

Python version: 2.7.11
Operating Systems tested: Windows 8.1 Pro 64 Bit, Windows 10 Home 64 Bit
*****************************************************************************'''

import xlrd     # for reading in xls and xlsx files
import csv      # for reading in csv files
import openpyxl # for reading and writing in xlsx and xlsm files
from openpyxl import load_workbook # Importing entire openpyxl didn't get me this. idk y...
import codecs   # for opening unicode files
import os       # for checking and manipulating file paths
import glob     # for finding files without knowing thier exact names
import shutil   # for moving files to different directories
import time     # for printing time stamps
import bs4
from bs4 import BeautifulSoup
import re
import sys

AsiteExstension = ".xls"
BimExstension = ".csv"
OutputExstension = ".xlsx"

def main():
    #*************************Read in Settings File*****************************    
    settingLines = readInSettingsFile()         
    
    #**********************readInAsiteDataSettings**************************
    inputAFilePath, inputAStartRow, inputAIDCol, inputACompanyCol, inputAStatusCol = readInAsiteDataSettings(settingLines,AsiteExstension)
    
    #**********************readInBimDataSettings****************************
    inputBFilePath, inputBStartRow, inputBIDCol, inputBCompanyCol, inputBStatusCol = readInBimDataSettings(settingLines,BimExstension)
    
    #***************Read in status translation Data Settings********************    
    BToAStatTransFileName = readInSetting(
        settingLines[15],"BIM_TO_ASITE_STATUS_TRANSLATIONS_PATH:")
    if not os.path.exists(BToAStatTransFileName): 
	pathError(BToAStatTransFileName,"BIM_TO_ASITE_STATUS_TRANSLATIONS:")
    
    INPUT = codecs.open(BToAStatTransFileName, encoding='utf-8')
    statTransLines = INPUT.readlines()
    INPUT.close()
    fileEmpty(statTransLines,BToAStatTransFileName)    
    #BToAStatTransFileName = checkTransFile(BToAStatTransFileName)    
    
    TranslationDelimiter = readInSetting(settingLines[16],
                                         "Translation_Delimiter:")      
    
    #*********************Read in output Data Settings************************** 
    dumpPath =  readInSetting(settingLines[20],"DUMP_FOLDER:")
    dumpPath = doesDumpPathExist(dumpPath)
	
    outputFileName = readInSetting(settingLines[19],"OUTPUT_PATH:")    
    outputFileName = doesOldOutputExist(outputFileName,".xlsx",dumpPath)
	
    #***********************Find and Print Data Analysis************************
    ACompanies, AStatuses, AIDs = readInHTML_XLS(inputAFilePath, dumpPath, inputAIDCol,inputACompanyCol,inputAStatusCol, inputAStartRow, "ASITE_INPUT_DIR:")    
	
    BCompanies, BStatuses, BIDs = readInCSV(inputBFilePath, dumpPath, inputBIDCol, inputBCompanyCol, inputBStatusCol, inputBStartRow, "BIM_INPUT_DIR:")    
      
    BToAStatuses = translateBimStatToAsiteStat(BToAStatTransFileName,
                                               TranslationDelimiter, BStatuses) 
    
    genXlsxOutPut(outputFileName,AIDs,BIDs,ACompanies,BCompanies,AStatuses,
                  BStatuses,BToAStatuses)


'''*****************************readInSettingsFile()****************************
PURPOSE: To read in the settings.txt

INPUT: None

OUTPUT: 'settingLines' the list of lines from the settings.txt in the form of 
        an arry of strings.

NOTES: If unable to find settings.txt an error message will be printed saying 
       so. The same goes for if the file is empty.
*****************************************************************************'''
def readInSettingsFile():
    if len(sys.argv)>1:
	for thing in sys.argv:
	    if thing[-len(".py"):] == ".py":
		thing = thing.replace("\\", "/")
		thing = thing.replace("//", "/")
		pos = recursivePath(thing,-1)
		outputFilepath = thing[:pos]
		if outputFilepath[1] == ":":
		    os.chdir(outputFilepath) 
		else:
		    print ("ERROR: "+outputFilepath+" is not a valid path.")
		    print ("Closing all File Explorer windows and try again.")		    
		    raw_input("\nEnter any key to EXIT program...")
		    sys.exit()
    settingsFile = "Data/settings.txt"
    if not os.path.exists(settingsFile):
	print ("ERROR: Cannot find the settings.txt file.")
	print ("This program cannot run properly without it.")
	# make the program stop so user acknowledges the failure to open
	junk = input("\nEnter any key to EXIT program...")
	sys.exit()     
    INPUT = codecs.open(settingsFile, encoding='utf-8')
    settingLines = INPUT.readlines()
    INPUT.close()
    fileEmpty(settingLines,settingsFile)
    
    return settingLines


'''**************************readInAsiteDataSettings()**************************
PURPOSE: To read in the settings from settings.txt required to properly read in
         the data from Asite.

INPUT: 'settingLines' the array of lines(strings) read in from the settings.txt.

OUTPUT: 'inputAFilePath' the file path for the file containing the required 
	data(in the form of a string), 'inputAStartRow' the number of the row 
	where the data starts(as an int), 'inputAIDCol' the number of the column 
	containing the company IDs(as an int), 'inputACompanyCol' the number of 
	the column containing the company names(as an int), and 
	'inputAStatusCol' the number of the column containing the company 
	statuses(as an int).
	
NOTES: None
*****************************************************************************'''
def readInAsiteDataSettings(settingLines, exstension):
    draged=False
    for thing in sys.argv:		
	if thing[-len(exstension):] == exstension:
	    #input("\nEnter any key to EXIT program...")
	    draged=True	
	    thing = thing.replace("\\", "/")
	    thing = thing.replace("//", "/")	    
	    inputAFilePath = thing
    if draged:
	for thing in sys.argv:
	    if thing[-len(".py"):] == ".py":
		thing = thing.replace("\\", "/")
		thing = thing.replace("//", "/")	    
		pos = recursivePath(thing,-1)
		outputFilepath = thing[:pos]
		os.chdir(outputFilepath)
    if not draged:
	inputAFilePath = readInSetting(settingLines[1],"ASITE_INPUT_DIR:")
    
    inputAStartRow = int(readInSetting(settingLines[2],
                                       "Asite_Header_Row_Num:"))-1
    inputAIDCol = int(readInSetting(settingLines[3],"Asite_ID_Col_Num:"))-1
    inputACompanyCol = int(readInSetting(settingLines[4],"Asite_Company_Col_Num:"))-1
    inputAStatusCol = int(readInSetting(settingLines[5],"Asite_Status_Col_Num:"))-1 
    
    
    return inputAFilePath, inputAStartRow, inputAIDCol, inputACompanyCol, inputAStatusCol

'''***************************readInBimDataSettings()***************************
PURPOSE: To read in the settings from settings.txt required to properly read in
         the data from BIM.

INPUT: 'settingLines' the array of lines(strings) read in from the settings.txt

OUTPUT: 'inputBFilePath' the file path for the file containing the required 
        data(in the form of a string), 'inputBStartRow' the number of the row 
	where the data starts(as an int), 'inputBCodeCol' the number of the 
	column containing the company IDs(as an int), 'inputBCompanyCol' the 
	number of the column containing the company Names(as an int), and 
	'inputBStatusCol' the number of the column containing the company 
	statuses(as an int).
	
NOTES: None
*****************************************************************************'''
def readInBimDataSettings(settingLines, exstension):
    draged=False
    for thing in sys.argv:		
	if thing[-len(exstension):] == exstension:
	    #input("\nEnter any key to EXIT program...")
	    draged=True	
	    thing = thing.replace("\\", "/")
	    thing = thing.replace("//", "/")	    
	    inputBFilePath = thing
    if draged:
	for thing in sys.argv:
	    if thing[-len(".py"):] == ".py":
		thing = thing.replace("\\", "/")
		thing = thing.replace("//", "/")	    
		pos = recursivePath(thing,-1)
		outputFilepath = thing[:pos]
		os.chdir(outputFilepath)    
    if not draged:
	inputBFilePath = readInSetting(settingLines[8],"BIM_INPUT_DIR:")
	
    inputBStartRow = int(readInSetting(settingLines[9],"BIM_Header_Row_Num:"))-1
    inputBIDCol  = int(readInSetting(settingLines[10],"BIM_ID_Col_Num:"))-1
    inputBCompanyCol = int(readInSetting(settingLines[11],"BIM_Company_Col_Num:"))-1
    inputBStatusCol = int(readInSetting(settingLines[12],"BIM_Status_Col_Num:"))-1    
    
    return inputBFilePath, inputBStartRow, inputBIDCol, inputBCompanyCol, inputBStatusCol

'''*******************************readInSetting()*******************************
PURPOSE: To read in a single setting from settings.txt required to properly 
         read in the required data.

INPUT: Two strings. 'line' a string read in from the settings.txt containing the 
       required setting and 'string' a string indicating what setting is being 
       read in.

OUTPUT: 'setting' a string containing only the required setting.
	
NOTES: If 'string' can not be found in 'line' the function will print an Error 
       message saying that the required line is either missing or entered wrong
       in the settings.txt.
*****************************************************************************'''
def readInSetting(line, string):
    if (line.find(string) != -1):
	setting = readInPromptEntries(line)
    else:
	lineMissingError(string)
    
    setting = setting.replace("\\", "/")
    setting = setting.replace("//", "/")    
    return setting

'''****************************readInPromptEntries()****************************
PURPOSE: To read in a single entry from a string after a ":". 

INPUT: 'nextLine' a string

OUTPUT: 'nextEntry' a string
	
NOTES: If there is a space after the ":" it will keep checking the next 
       character untill it finds one that isn't white space or untill it reaches 
       the end of the string. If unable to find a ":" will just return the 
       original string.
*****************************************************************************'''
def readInPromptEntries(nextLine):
    nextLine = nextLine.strip()
    if (nextLine.find(": ") != -1):
	offset = recursiveSpaces(nextLine,nextLine.find(":")+1,1)
    else:
	offset = 0
    entryStart = nextLine.find(":")
    nextEntry = nextLine[(entryStart+1+offset):]
    return nextEntry

'''******************************recursiveSpaces()******************************
PURPOSE: To find out how many spaces there are in a row after the given position
         in a string. 

INPUT: 'line' a string, 'pos' an int conaining a numeric position in that 
       string, and 'howMany' an int containing number of spaces to start form in 
       the count.

OUTPUT: 'howMany' an int containing the number of spaces found plus the starting 
	number of spaces 
	
NOTES: None
*****************************************************************************'''
def recursiveSpaces(line,pos,howMany):
    if line[pos+1]==' ':
	howMany = recursiveSpaces(line,pos+1,howMany+1)
    return howMany

'''********************************readInXLSM()*********************************
PURPOSE: Use the setting found in the setting.txt to read in the required data 
	 from a .xlsm file

INPUT: 'inputFilePath' a string containing a filepath(with or with out a file 
       name at the end of it) to where a .xlsm file can be found, 'dumpPath' a 
       filepath to a folder to dump used or old files, 'inputIDCol' the column 
       number for the company IDs, 'inputCompanyCol' the column number for the 
       company names,'inputStatusCol' the column number for the company 
       statuses, 'inputStartRow' the row nuber for where the data starts, and 
       'identifier' a string containing the identifier for the required 
       file(i.e. 'my 3rd grade report on dogs').

OUTPUT: 'Companies' a list of strings containing the company names, 'Statuses' a 
        list containing the company statuses, and 'IDs' a list containing the 
	company IDs.
	
NOTES: Will rename the file that the data was read in from with a time stamp and
       move it to the dump folder afterwords.
*****************************************************************************'''
def readInXLSM(inputFilePath,dumpPath,inputIDCol,inputCompanyCol,inputStatusCol,
               inputStartRow, identifier):
    exstension = ".xlsm"
    inputFilePath, inputFileName = IdentifyFileAndPath(inputFilePath, 
                                                       exstension, identifier)
    
    wb = load_workbook(inputFileName)
    Company =  wb.active    
    
    Companies = []
    Statuses = []
    IDs = []         
    
    for i in range(Company.max_row):
	if i>inputStartRow:
	    Companies.append(Company.cell(row = i, 
	                                  column = (inputCompanyCol+1)).value)
	    Statuses.append(Company.cell(row = i,
	                                 column = (inputStatusCol+1)).value)
	    IDs.append(Company.cell(row = i, column = (inputIDCol+1)).value)
    
    checkDumpForDoubles(inputFileName,inputFilePath,dumpPath,"AsiteDir",
                        exstension)     
    
    return Companies, Statuses, IDs

'''********************************readInXLSX()*********************************
PURPOSE: Use the setting found in the setting.txt to read in the required data 
	 from a .xlsx file

INPUT: 'inputFilePath' a string containing a filepath(with or with out a file 
       name at the end of it) to where a .xlsx file can be found, 'dumpPath' a 
       filepath to a folder to dump used or old files, 'inputIDCol' the column 
       number for the company IDs, 'inputCompanyCol' the column number for the 
       company names,'inputStatusCol' the column number for the company 
       statuses, 'inputStartRow' the row nuber for where the data starts, and 
       'identifier' a string containing the identifier for the required 
       file(i.e. 'my 3rd grade report on dogs').

OUTPUT: 'Companies' a list of strings containing the company names, 'Statuses' a 
        list containing the company statuses, and 'IDs' a list containing the 
	company IDs.
	
NOTES: Will rename the file that the data was read in from with a time stamp and
       move it to the dump folder afterwords.
*****************************************************************************'''
def readInXLSX(inputFilePath,dumpPath,inputIDCol,inputCompanyCol,inputStatusCol,
               inputStartRow, identifier):
    exstension = ".xlsx"
    inputFilePath, inputFileName = IdentifyFileAndPath(inputFilePath, 
                                                       exstension, identifier)
    
    wb = load_workbook(inputFileName)
    Company =  wb.active    
    
    Companies = []
    Statuses = []
    IDs = []         
    
    for i in range(Company.max_row):
	if i>inputStartRow:
	    Companies.append(Company.cell(row = i, 
	                                  column = (inputCompanyCol+1)).value)
	    Statuses.append(Company.cell(row = i, 
	                                 column = (inputStatusCol+1)).value)
	    IDs.append(Company.cell(row = i, column = (inputIDCol+1)).value)
    
    checkDumpForDoubles(inputFileName,inputFilePath,dumpPath,"AsiteDir",
                        exstension)     
    
    return Companies, Statuses, IDs

'''*********************************readInXLS()*********************************
PURPOSE: Use the setting found in the setting.txt to read in the required data 
	 from a .xls file

INPUT: 'inputFilePath' a string containing a filepath(with or with out a file 
       name at the end of it) to where a .xls file can be found, 'dumpPath' a 
       filepath to a folder to dump used or old files, 'inputIDCol' the column 
       number for the company IDs, 'inputCompanyCol' the column number for the 
       company names,'inputStatusCol' the column number for the company 
       statuses, 'inputStartRow' the row nuber for where the data starts, and 
       'identifier' a string containing the identifier for the required 
       file(i.e. 'my 3rd grade report on dogs').

OUTPUT: 'Companies' a list of strings containing the company names, 'Statuses' a 
        list containing the company statuses, and 'IDs' a list containing the 
	company IDs.
	
NOTES: Will rename the file that the data was read in from with a time stamp and
       move it to the dump folder afterwords.
*****************************************************************************'''
def readInXLS(inputFilePath,dumpPath,inputIDCol,inputCompanyCol,inputStatusCol,
              inputStartRow, identifier):
    exstension = ".xls"
    inputFilePath, inputFileName = IdentifyFileAndPath(inputFilePath, 
                                                       exstension, identifier)
    
    Company = xlrd.open_workbook(inputFileName).sheet_by_index(0)
	
    Companies = Company.col_values(inputCompanyCol, start_rowx=inputStartRow)
    Statuses = Company.col_values(inputStatusCol, start_rowx=inputStartRow)
    IDs = Company.col_values(inputIDCol, start_rowx=inputStartRow)
    
    checkDumpForDoubles(inputFileName,inputFilePath,dumpPath,"AsiteDir",
                        exstension)     
    
    return Companies, Statuses, IDs

'''*********************************readInCSV()*********************************
PURPOSE: Use the setting found in the setting.txt to read in the required data 
	 from a .csv file

INPUT: 'inputFilePath' a string containing a filepath(with or with out a file 
       name at the end of it) to where a .csv file can be found, 'dumpPath' a 
       filepath to a folder to dump used or old files, 'inputIDCol' the column 
       number for the company IDs, 'inputCompanyCol' the column number for the 
       company names,'inputStatusCol' the column number for the company 
       statuses, 'inputStartRow' the row nuber for where the data starts, and 
       'identifier' a string containing the identifier for the required 
       file(i.e. 'my 3rd grade report on dogs').

OUTPUT: 'Companies' a list of strings containing the company names, 'Statuses' a 
        list containing the company statuses, and 'IDs' a list containing the 
	company IDs.
	
NOTES: Will rename the file that the data was read in from with a time stamp and
       move it to the dump folder afterwords.
*****************************************************************************'''
def readInCSV(inputFilePath, dumpPath,inputIDCol,inputCompanyCol,inputStatusCol, inputStartRow, identifier):
    exstension = ".csv"
    inputFilePath, inputFileName = IdentifyFileAndPath(inputFilePath, 
                                                       exstension, identifier)   
    Companies = []
    Statuses = []
    IDs = []
    with open(inputFileName, 'rb') as csvfile:
	BINPUT = csv.reader(csvfile)
	
	if inputStartRow != -1:
	    i=0
	    while(i<(inputStartRow)):
		i=i+1
		headers = BINPUT.next()
	else:
	    i=0
	    while inputStartRow == -1:
		headers = BINPUT.next()
		if ("Description" in headers) or ("Name" in headers) or ("Status" in headers):
		    inputStartRow = i
		else: i=i+1
		
	if ((inputIDCol == -1) and ("Description" in headers)):
	    inputIDCol =(headers.index("Description"))
	elif ((inputIDCol == -1) and (not("Description" in headers))):
	    ColNotFound("Description")
	    
	if ((inputCompanyCol == -1) and ("Name" in headers)):
	    inputCompanyCol=(headers.index("Name"))
	elif ((inputCompanyCol == -1) and (not("Name" in headers))):
	    ColNotFound("Name")	
	    
	if ((inputStatusCol == -1) and ("Status" in headers)):
	    inputStatusCol=(headers.index("Status"))
	elif ((inputStatusCol == -1) and (not("Status" in headers))):
	    ColNotFound("Status")	
	    
	for nextRow in BINPUT:
	    Companies.append(nextRow[inputCompanyCol])
	    IDs.append(nextRow[inputIDCol])
	    Statuses.append(nextRow[inputStatusCol]) 
	    
    checkDumpForDoubles(inputFileName,inputFilePath,dumpPath, "BimDir",
                        exstension)    
    
    return Companies, Statuses, IDs

'''*********************************readInHTML_XLS()*********************************
PURPOSE: Use the setting found in the setting.txt to read in the required data 
	 from a file containing HTML code but has a .xls extension.

INPUT: 'inputFilePath' a string containing a filepath(with or with out a file 
       name at the end of it) to where a .xls file can be found, 'dumpPath' a 
       filepath to a folder to dump used or old files, 'inputIDCol' the column 
       number for the company IDs, 'inputCompanyCol' the column number for the 
       company names,'inputStatusCol' the column number for the company 
       statuses, 'inputStartRow' the row nuber for where the data starts, and 
       'identifier' a string containing the identifier for the required 
       file(i.e. 'my 3rd grade report on dogs').

OUTPUT: 'Companies' a list of strings containing the company names, 'Statuses' a 
        list containing the company statuses, and 'IDs' a list containing the 
	company IDs.
	
NOTES: Will rename the file that the data was read in from with a time stamp and
       move it to the dump folder afterwords.
*****************************************************************************'''
def readInHTML_XLS(inputFilePath, dumpPath, inputIDCol,inputCompanyCol,inputStatusCol, inputStartRow, identifier):
    exstension = ".xls"
    inputFilePath, inputFileName = IdentifyFileAndPath(inputFilePath, 
                                                       exstension, identifier)    
    f = open(inputFileName)
    soup = BeautifulSoup(f, "html.parser")
    f.close()
    Companies = []
    Statuses = []
    IDs = []
    inputRow=0
    startingRows = soup.find_all('tr')
    while inputStartRow == -1:
	header = soup.find_all('tr')[inputRow]
	headers = header.find_all(text=True)
	headers=remAll(headers,"\n")
	if ("ID" in headers) or ("Form Title" in headers) or ("Status" in headers):
	    inputStartRow=inputRow
	else: inputRow=inputRow+1
	
    if ((inputIDCol == -1) and ("ID" in headers)):
	inputIDCol=(headers.index("ID"))#.decode('unicode-escape')))
    elif ((inputIDCol == -1) and (not("ID" in headers))):
	ColNotFound("ID")    
	
    if ((inputCompanyCol == -1) and ("Form Title" in headers)):
	inputCompanyCol=(headers.index("Form Title"))
    elif ((inputCompanyCol == -1) and (not("Form Title" in headers))):
	ColNotFound("Form Title")
	
    if ((inputStatusCol == -1) and ("Status" in headers)):
	inputStatusCol=(headers.index("Status"))    
    elif ((inputStatusCol == -1) and (not("Status" in headers))):
	ColNotFound("Status")
	
    rows = soup.find_all('tr')[(inputStartRow+1):]
    for row in rows:
	cells = row.find_all('td')
	if not((len(cells)<inputCompanyCol) or (len(cells)<inputIDCol) or (len(cells)<inputStatusCol)):
	    Companies.append(cells[inputCompanyCol].find(text=True))
	    IDs.append(cells[inputIDCol].find(text=True))
	    Statuses.append(cells[inputStatusCol].find(text=True))
	    
    checkDumpForDoubles(inputFileName,inputFilePath,dumpPath, "AsiteDir",
                        exstension)    
    
    return Companies, Statuses, IDs   

def remAll(L, item):
    answer = []
    for i in L:
        if i!=item:
            answer.append(i)
    return answer

'''****************************IdentifyFileAndPath()****************************
PURPOSE: Find both the path of the folder that the required file is in(path 
         without the file name at the end) and the path to the required 
	 file(path with the file name at the end) using a file path that does or
	 doesn't have a file name at the end of it.

INPUT: 'inputFilePath' a string containing a filepath(with or with out a file 
       name at the end of it) to where a file with the given extension can be 
       found, 'exstension' a string containing the extension(i.e. '.py') of the 
       required file, and 'identifier' a string containing the identifier for 
       the required file(i.e. 'my 3rd grade report on dogs'). 
 
OUTPUT: 'inputFilePath' a string containing the path to the folder where a file 
	with the given extension can be found(path without a file name at the 
	end) and 'inputFileName' a string containing the path to a file with the 
	given extension. 
	
NOTES: The filepath is expected to use single '/' to indicate pathing. If the 
       identifier string is "OUTPUT_PATH:" and the filepath input string is 
       empty an error message get printed expliaing that line in the 
       settings.txt should have a file name to generate the file with.
*****************************************************************************'''
def IdentifyFileAndPath(inputFilePath,exstension,identifier):
    inputFileName = "NULL"
    if len(inputFilePath) > 3:
	if (inputFilePath[-len(exstension):] == exstension):
	    pos = recursivePath(inputFilePath,-len(exstension))
	    if pos != -1:
		inputFileName = inputFilePath
		inputFilePath = inputFilePath[:pos]
	    else:
		inputFileName = inputFilePath
		inputFilePath = ""		
	else:
	    if  not ("/" in inputFilePath[-1]): 
		inputFilePath = inputFilePath + "/"
	    if len(glob.glob(os.path.join(inputFilePath,'*'+exstension)))>0:
		inputFileName = glob.glob(os.path.join(inputFilePath,
		                                       '*'+exstension))[0]
    elif len(inputFilePath)<1: 
	if identifier == "OUTPUT_PATH:":
	    print ("ERROR: The line \"OUTPUT_PATH:\" in settings has no file path to write to.")
	    print ("This program cannot run properly without it.")
	    print ("Please check that the file path is entered correctly in the settings.txt file.")
	    print ("Consult the ReadMe.txt if you are not sure of how to properly modify the file.")
	    # make the program stop so user acknowledges the failure to open
	    junk = input("\nEnter any key to EXIT program...")
	    sys.exit()	    
	elif len(glob.glob('*'+exstension))>0: 
	    inputFileName = glob.glob('*'+exstension)[0]
    if not (os.path.exists(inputFileName) or (identifier == "OUTPUT_PATH:")):
	pathError(inputFilePath,identifier)
	
    return inputFilePath, inputFileName
    
'''*********************************recursivePath()*****************************
PURPOSE: To find the first leftmost '/' from a perticular position in a string.

INPUT: 'line' a string and 'pos' a int containg the index of the string to start
       at. 
 
OUTPUT: 'pos' an int containing the index of the first leftmost '/' from the 
	given position.
	
NOTES: If no '/' is found -1 will be returned. 
*****************************************************************************'''
def recursivePath(line,pos):
    if "/" in line:
	if  not ("/" in line[pos-1]):
	    pos = recursivePath(line,pos-1)
    else: pos = -1
    return pos

'''****************************checkDumpForDoubles()****************************
PURPOSE: Renames the given file with the given name and puts a timestamp at the 
         end of it and then checks if the new filename already exsists in the 
	 dump folder. If it does a number gets added at the end to make it 
	 unquie and then gets put in the dump folder. If it does not it is 
	 simply moved to the dump folder.

INPUT: 'inputFileName' a string containg a filepath, 'inputFilePath' a string 
       containing the path to the folder that that file is in, 'dumpPath' a 
       string containg the path to the dump folder, 'sourceDir' a string the 
       containing the name to be given to the file, 'exstension' a string 
       containing the extension of the file.
 
OUTPUT: None
	
NOTES: If 'sourceDir' is equal 'exstension' the file will not be renamed; only 
       moved.
*****************************************************************************'''
def checkDumpForDoubles(inputFileName,inputFilePath,dumpPath, sourceDir,exstension):
    exLen=len(exstension)
    currentFileName = inputFileName[len(inputFilePath):]
    if ((os.path.join(inputFilePath,currentFileName) != 
        os.path.join(dumpPath,currentFileName)) or 
        (sourceDir == exstension)):
	
	if sourceDir != exstension:
	    currentFileName = time.strftime(sourceDir+"-%y%m%d-%H%M%S"
	                                    +currentFileName[-exLen:])
	    os.rename(inputFileName,os.path.join(inputFilePath,currentFileName))
	inputFileName = os.path.join(inputFilePath,currentFileName)
	if (os.path.exists(os.path.join(dumpPath,currentFileName))):
	    updatedFileName = findFileNameDouble(dumpPath,currentFileName,
	                                         0,exLen)
	    os.rename(inputFileName,os.path.join(inputFilePath,updatedFileName))
	    inputFileName = os.path.join(inputFilePath,updatedFileName)	
	shutil.move(inputFileName, os.path.join(dumpPath,currentFileName))    

'''****************************findFileNameDouble()*****************************
PURPOSE: Find how many other doubles there are of the given file name and add 
         the number plus one to the end of the filename to name the new double 
	 with.

INPUT: 'dumpPath' a string containg the path to the dump folder, 'FileName' 
       string containg a filename, 'count' an int containg the number of times 
       this function has been run on the current parameters(should be given 0 to 
       start or 1 if the filename string is already a result of this function), 
       'exLen' an int containing the length of the extension of the file.
 
OUTPUT: 'FileName' a string containing a filename that does not yet exsist in 
	the dump folder.
	
NOTES: None
*****************************************************************************'''
def findFileNameDouble(dumpPath,FileName,count,exLen):
    if os.path.exists(os.path.join(dumpPath,FileName)):
	if(count>0):
	    FileName=FileName[:-(len(str(count+1))+7)]+FileName[-exLen:]
	count=count+1
	FileName = findFileNameDouble(dumpPath,FileName[:-exLen]+" ("+
	                              str(count+1)+")"+FileName[-exLen:],
	                              count,exLen)
    
    return FileName    

'''*****************************doesDumpPathExist()*****************************
PURPOSE: Determin whether the given dump folder path actually exsists.

INPUT: 'dumpPath' a string containg the path to the dump folder.
 
OUTPUT: 'dumpPath' a string containg the path to the dump folder.
	
NOTES: If 'dumpPath' does not contain a folder path that exsists then an messege
       will be printed saying so.
*****************************************************************************'''
def doesDumpPathExist(dumpPath):
    if len(dumpPath) > 0:
	if  not ("/" in dumpPath[-1]): 
	    dumpPath = dumpPath + "/"
	if not os.path.exists(dumpPath): pathError(dumpPath,"DUMP_FOLDER:")
    return dumpPath

'''****************************doesOldOutputExist()*****************************
PURPOSE: Add a time stamp to the filename of the new output file. If another 
         file with the same extension as the output file exsists in the output 
	 directory move it to the dump folder and rename it if nessisary. 

INPUT: 'outputFileName' a string containing the filepath and name and extension 
       of the outputfile that needs to be generated, 'extension' a string 
       containing just the extension for the output file, 'dumpPath' a string 
       containg the path to the dump folder.
 
OUTPUT: 'outputFileName' a string containing the filepath and name and time 
        stamp and extension of the outputfile that needs to be generated.
	
NOTES: None
*****************************************************************************'''
def doesOldOutputExist(outputFileName,extension,dumpPath):
    exLen = len(extension)
    outputFilePath, outputFileName = IdentifyFileAndPath(outputFileName,
                                                         extension,
                                                         "OUTPUT_PATH:")
    if ((len(glob.glob(os.path.join(outputFilePath,'*'+extension)))>0) 
            and (outputFilePath != dumpPath)):
	lastOutputFileName = glob.glob(os.path.join(outputFilePath,
                                                    '*'+extension))[0]
	sourceDir = extension
	checkDumpForDoubles(lastOutputFileName,outputFilePath,dumpPath, 
                            sourceDir,extension)
    currentFileName = outputFileName[len(outputFilePath):] 
    currentFileName = time.strftime(currentFileName[:-exLen]
                                    +"-%y%m%d-%H%M%S"+currentFileName[-exLen:])
    outputFileName = os.path.join(outputFilePath,currentFileName)
    
    return outputFileName

'''************************translateBimStatToAsiteStat()************************
PURPOSE: To translate the Bim statuses to the Asite statuses so they can be 
         compared to the Asite data.

INPUT: 'BToAStatTransFileName' a string containing the filepath of the file 
       containing the translations for Bim statuses into Asite statuses, 
       'TranslationDelimiter' a character that contains the delimiter for Bim 
       statuses in the translation file, 'BStatuses' a string containing all the
       Bim statuses from the Bim input data.
 
OUTPUT: 'BToAStatuses' a list of strings containing all the Bim statuses from 
	the Bim input data translated into their respective Asite equivalants.
	
NOTES: None
*****************************************************************************'''
def translateBimStatToAsiteStat(BToAStatTransFileName,TranslationDelimiter,
                                BStatuses):
    INPUT = codecs.open(BToAStatTransFileName, encoding='utf-8')
    statTransLines = INPUT.readlines()
    INPUT.close()
    
    statTransList = []
    statList = []
    for nextKeyWords in statTransLines:
	statTransList.append(
	    readInPromptEntries(nextKeyWords).split(TranslationDelimiter))
	statList.append(nextKeyWords[:(nextKeyWords.find(":"))])
    
    BToAStatuses = []
    for nextStat in BStatuses:
	for j in range(len(statTransList)):
	    for i in range(len(statTransList[j])):
		if statTransList[j][i][0] == ' ':
		    offset = recursiveSpaces(statTransList[j][i],0,1)
		    statTransList[j][i] = statTransList[j][i][offset:]
		if (nextStat.find(statTransList[j][i]) != -1):
		    nextStat = nextStat.replace(nextStat,statList[j]) 
	BToAStatuses.append(nextStat) 
	
    return BToAStatuses

''' *******************************genTsvOutPut()*******************************
PURPOSE: To analyze the Bim and Asite data and the print the analysis in the 
         form of a .tsv file.

INPUT: 'outputFileName' a string containing the filepath and name and extension 
       of the output file, 'AIDs' list a strings that that contains the company 
       IDs from the Asite database, 'BIDs' list a strings that that contains the 
       company IDs from the Bim database, 'ACompanies' a list of strings that 
       contains the company names from the Asite database, 'BCompanies' a list 
       of strings that contains the company names from the Bim database, 
       'AStatuses' a list of strings containing the statuses from the Asite 
       database, 'BStatuses' a list of strings containing the statuses from the 
       Bim database, 'BToAStatuses' a list of string containing the statuses 
       from the Bim database translated into their respective Asite equivalants.
 
OUTPUT: A .tsv file containing a comparison between the Asite and Bim databases.
	
NOTES: Prompts user to acknowledge that the program is finished running at the 
       end of this function.
*****************************************************************************'''
def genTsvOutPut(outputFileName,AIDs,BIDs,ACompanies,BCompanies,AStatuses,
                 BStatuses,BToAStatuses):
    true="Yes"
    false="No"
    outputFile = codecs.open(outputFileName, encoding='utf-8', mode='w')
    outputFile.write("Asite ID\tBIM ID\tMissing\tAsite Company\tBIM Company\tNaming Discrepancy\tAsite Status\tBIM Status\tStatus Discrepancy\tHas BIM Double\n")
    for i in range(len(AIDs)):
	isMissing=true
	for j in range(len(BIDs)):
	    nameDiscrepancy = true
	    statusDiscrepancy = true
	    isDouble=false	    
	    if (AIDs[i] == BIDs[j]):
		isMissing=false
		if BIDs.count(BIDs[j]) > 1:
		    isDouble=true
		
		if (ACompanies[i] == BCompanies[j]):
		    nameDiscrepancy = false
		    
		if (AStatuses[i] == BToAStatuses[j]):
		    statusDiscrepancy = false
		    
		outputFile.write("%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\n" % 
		                 (AIDs[i], BIDs[j], isMissing, ACompanies[i], 
		                  BCompanies[j], nameDiscrepancy, AStatuses[i], 
		                  BStatuses[j], statusDiscrepancy,isDouble))
			
	if (isMissing==true):
	    outputFile.write("%s\t \tOn BIM\t%s\t \t--\t%s\t \t--\t--\n" % 
	                     (AIDs[i], ACompanies[i], AStatuses[i]))
	
    for i in range(len(BIDs)):
	isMissing=True
	for j in range(len(AIDs)):
	    if (BIDs[i] == AIDs[j]):
		isMissing=False
		
	if (isMissing==True):
	    outputFile.write(" \t%s\tOn Asite\t \t%s\t--\t \t%s\t--\t--\n" % 
	                     (BIDs[i], BCompanies[i], BStatuses[i]))	

    outputFile.close()
    
    print("Operation Complete!!!")
    junk = input("\nEnter any key to EXIT program.")
    sys.exit()    

'''*******************************genXlsxOutPut()*******************************
PURPOSE: To analyze the Bim and Asite data and the print the analysis in the 
         form of a .xlsx or .xlsm file.

INPUT: 'outputFileName' a string containing the filepath and name and extension 
       of the output file, 'AIDs' list a strings that that contains the company 
       IDs from the Asite database, 'BIDs' list a strings that that contains the 
       company IDs from the Bim database, 'ACompanies' a list of strings that 
       contains the company names from the Asite database, 'BCompanies' a list 
       of strings that contains the company names from the Bim database, 
       'AStatuses' a list of strings containing the statuses from the Asite 
       database, 'BStatuses' a list of strings containing the statuses from the 
       Bim database, 'BToAStatuses' a list of string containing the statuses 
       from the Bim database translated into their respective Asite equivalants.
 
OUTPUT: A .xlsx or .xlsm file containing a comparison between the Asite and Bim 
        databases.
	
NOTES: Prompts user to acknowledge that the program is finished running at the 
       end of this function.
*****************************************************************************'''
def genXlsxOutPut(outputFileName,AIDs,BIDs,ACompanies,BCompanies,AStatuses,
                  BStatuses,BToAStatuses):
    true="Yes"
    false="No"    
    output = openpyxl.Workbook(optimized_write=True)
    outRow = 0
    outputFile = output.create_sheet(0)
    cols=["Asite ID","BIM ID","Missing","Asite Company","BIM Company",
          "Naming Discrepancy","Asite Status","BIM Status","Status Discrepancy",
          "Has BIM Double"] 
    outputFile.append(cols)   
    
    for i in range(len(AIDs)):
	isMissing=true
	outRow = outRow+1
	for j in range(len(BIDs)):
	    nameDiscrepancy = true
	    statusDiscrepancy = true
	    isDouble=false	    
	    if (AIDs[i] == BIDs[j]):
		isMissing=false
		if BIDs.count(BIDs[j]) > 1:
		    isDouble=true
		if (ACompanies[i] == BCompanies[j]):
		    nameDiscrepancy = false  
		if (AStatuses[i] == BToAStatuses[j]):
		    statusDiscrepancy = false
		    
		outputFile.append([AIDs[i], BIDs[j], isMissing, ACompanies[i], 
		                   BCompanies[j], nameDiscrepancy, AStatuses[i],
		                   BStatuses[j], statusDiscrepancy,isDouble])			
			
	if (isMissing==true):
	    cols = [AIDs[i],"","On BIM", ACompanies[i],"","--", AStatuses[i],"",
	            "--","--"]
	    outputFile.append(cols)				    

    for i in range(len(BIDs)):
	outRow = outRow+1
	isMissing=True
	for j in range(len(AIDs)):
	    if (BIDs[i] == AIDs[j]):
		isMissing=False
		
	if (isMissing==True):
	    cols = ["",BIDs[i],"On Asite","", BCompanies[i],"--","", 
	            BStatuses[i],"--","--"]
	    outputFile.append(cols)     
	
    
    output.save(outputFileName)
    
    print("Operation Complete!!!")
    junk = input("\nEnter any key to EXIT program.")
    sys.exit()

def pathError(path, identifier):
    print ("ERROR: Cannot find the following "+identifier+ " " + path)
    print ("This program cannot run properly without it.")
    print ("Please check that the file path is entered correctly in the settings.txt file.")
    print ("Consult the ReadMe.txt if you are not sure of how to properly modify the file.")
    # make the program stop so user acknowledges the failure to open
    junk = input("\nEnter any key to EXIT program...")
    sys.exit()
	
def fileEmpty(lines, path):
    if not (len(lines)>0):
	print ("ERROR: "+path+" is empty!!!")
	print ("This program cannot run properly without it.")
	print ("Please check that required information is properly entered in this file.")
	print ("Consult the ReadMe.txt if you are not sure of how to properly modify the file.")
	# make the program stop so user acknowledges the failure to open
	junk = input("\nEnter any key to EXIT program...")
	sys.exit()
	
def lineMissingError(string):
    print ("ERROR: the line "+string+" is missing or not where it's supposed to be in settings.txt!!!")
    print ("This program cannot run properly without it.")
    print ("Please check that required information is properly entered in this file.")
    print ("Consult the ReadMe.txt if you are not sure of how to properly modify the file.")
    # make the program stop so user acknowledges the failure to open
    junk = input("\nEnter any key to EXIT program...")
    sys.exit()

def ColNotFound(string):
    print ("ERROR: Could not find the "+string+" column in one of the input files!!!")
    print ("This program cannot run properly without it.")
    print ("To fix this manually enter the appropriate column number in the settings.txt")
    print ("Consult the ReadMe.txt if you are not sure of how to properly modify the file.")
    # make the program stop so user acknowledges the failure to open
    junk = input("\nEnter any key to EXIT program...")
    sys.exit()

main() #Written using python version 2.7.11 on Windows 8.1 Pro